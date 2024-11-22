from datetime import datetime
import json
import os
from typing import Dict, List
import pandas as pd
import matplotlib.pyplot as plt
from jira import JIRA
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pytz

class JiraBugStatistics:
    def __init__(self, config_path='config.json'):
        # 获取脚本所在目录
        script_dir = os.path.dirname(os.path.abspath(__file__))
        # 构建配置文件的完整路径
        config_path = os.path.join(script_dir, config_path)
        
        print(f"Looking for config file at: {config_path}")
        
        with open(config_path) as f:
            config = json.load(f)
            
        # 初始化JIRA客户端
        self.jira = JIRA(
            server=config['jira_server'],
            basic_auth=(config['username'], config['password'])
        )
        
        # 初始化缓存
        self.cache = {}
        
        # 添加输出目录属性
        self.output_dir = None
        
    def get_target_from_jql(self, jql: str) -> str:
        """从JQL中提取Target值"""
        import re
        target_match = re.search(r'Target\s*=\s*(\w+)', jql)
        return target_match.group(1) if target_match else 'Unknown'
    
    def create_output_directory(self, target: str) -> str:
        """创建输出目录"""
        from datetime import datetime
        import pytz
        
        # 获取当前中国时间
        timestamp = datetime.now(pytz.timezone('Asia/Shanghai')).strftime("%Y%m%d_%H%M%S")
        
        # 创建基础输出目录
        base_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Report')
        if not os.path.exists(base_dir):
            os.makedirs(base_dir)
            
        # 创建特定报告目录
        report_dir = os.path.join(base_dir, f"{target}_{timestamp}")
        if not os.path.exists(report_dir):
            os.makedirs(report_dir)
            
        return report_dir

    def get_issues(self, filter_id: str) -> List[Dict]:
        """获取所有issues"""
        try:
            # 首先获取过滤器的 JQL
            jql = self.jira.filter(filter_id).jql
            target = self.get_target_from_jql(jql)
            
            issues = []
            start_at = 0
            max_results = 50

            while True:
                batch = self.jira.search_issues(
                    jql,
                    startAt=start_at,
                    maxResults=max_results,
                    fields=['summary', 'status', 'customfield_11214', 'customfield_11219', 
                           'reporter', 'assignee', 'created', 'id']
                )
                
                if not batch:
                    break
                    
                issues.extend(batch)
                if len(batch) < max_results:
                    break
                    
                start_at += max_results
                
            return issues
        except Exception as e:
            print(f"Error fetching issues: {str(e)}")
            return []

    def analyze_issues(self, filter_id: str) -> pd.DataFrame:
        """分析issues并返回统计数据"""
        issues = self.get_issues(filter_id)
        
        # 初始化计数器字典
        stats = {
            'U0 Blocking': {'To Do': 0, 'In Progress': 0, 'Resolved': 0, 'Closed': 0},
            'U1 Urgent': {'To Do': 0, 'In Progress': 0, 'Resolved': 0, 'Closed': 0},
            'U2 Normal': {'To Do': 0, 'In Progress': 0, 'Resolved': 0, 'Closed': 0},
            'U3 Low': {'To Do': 0, 'In Progress': 0, 'Resolved': 0, 'Closed': 0},
            'None': {'To Do': 0, 'In Progress': 0, 'Resolved': 0, 'Closed': 0}
        }
        
        print(f"\nProcessing {len(issues)} issues...")
        
        for issue in issues:
            # 获取状态名称
            status = issue.fields.status.name
            
            # 获取 Urgency 值
            urgency = getattr(issue.fields, 'customfield_11214', None)
            if urgency:
                urgency = str(urgency)
                # 确保 urgency 是有效的类别
                if urgency not in stats:
                    print(f"Warning: Unexpected urgency '{urgency}' for issue {issue.key}")
                    urgency = 'None'
            else:
                urgency = 'None'
            
            print(f"Issue {issue.key}: Urgency = {urgency}, Status = {status}")  # 调试信息
            
            # 标准化状态名称
            if status not in ['To Do', 'In Progress', 'Resolved', 'Closed']:
                print(f"Warning: Unexpected status '{status}' for issue {issue.key}")
                continue
            
            # 更新统计
            stats[urgency][status] += 1
        
        # 创建DataFrame
        df = pd.DataFrame(stats).T
        
        # 确保所有状态列都存在
        for status in ['To Do', 'In Progress', 'Resolved', 'Closed']:
            if status not in df.columns:
                df[status] = 0
        
        # 添加行总计
        df['Total'] = df.sum(axis=1)
        
        # 添加列总计
        df.loc['Total'] = df.sum()
        
        # 打印调试信息
        print("\nFinal statistics:")
        print(df)
        
        return df

    def create_pie_chart(self, df):
        """创建饼图"""
        # 检查是否有数据
        if df.empty:
            print("No data available for pie chart")
            return
        
        # 定义状态和颜色的固定映射
        status_colors = {
            'To Do': '#f49513',
            'In Progress': '#f4f413',
            'Resolved': '#76ccf2',
            'Closed': '#58e790'
        }
        
        # 计算每个状态的总数
        status_counts = df[list(status_colors.keys())].sum()
        
        # 检查是否所有值都为0
        if status_counts.sum() == 0:
            print("No issues found in any status")
            return
        
        # 只保留有数据的状态
        status_counts = status_counts[status_counts > 0]
        
        # 获取对应的颜色列表，保持与状态的对应关系
        colors = [status_colors[status] for status in status_counts.index]
        
        # 如果还有数据，则创建饼图
        if not status_counts.empty:
            # 导入seaborn并设置样式
            import seaborn as sns
            sns.set_style("whitegrid")
            
            # 创建图形，设置更大的尺寸
            plt.figure(figsize=(12, 8))
            
            # 创建饼图
            wedges, texts, autotexts = plt.pie(
                status_counts,
                labels=status_counts.index,
                autopct='%1.1f%%',
                startangle=90,
                colors=colors,  # 使用对应的颜色列表
                pctdistance=0.85,
                explode=[0.05] * len(status_counts),
                wedgeprops={'edgecolor': 'white', 'linewidth': 2}
            )
            
            # 设置字体为 Calibri
            plt.rcParams['font.family'] = 'Calibri'
            
            # 设置标签和百分比的样式
            plt.setp(texts, size=11, family='Calibri', weight='bold')
            plt.setp(autotexts, size=10, family='Calibri', weight='bold')
            
            # 添加标题
            plt.title('Bug Status Distribution', 
                     pad=20,
                     size=14, 
                     family='Calibri',
                     weight='bold')
            
            # 添加图例，显示所有状态（包括数量为0的）
            legend_labels = [f'{status} ({int(status_counts.get(status, 0))})' 
                            for status in status_colors.keys()]
            legend_handles = [plt.Rectangle((0,0),1,1, fc=color) 
                             for color in status_colors.values()]
            
            plt.legend(legend_handles, 
                      legend_labels,
                      title="Status",
                      loc="center left",
                      bbox_to_anchor=(1, 0, 0.5, 1))
            
            # 确保图例不被裁切
            plt.tight_layout()
            
            # 修改保存路径
            pie_chart_path = os.path.join(self.output_dir, 'bug_status_pie_chart.png')
            plt.savefig(pie_chart_path, 
                       dpi=300, 
                       bbox_inches='tight',
                       facecolor='white',
                       edgecolor='none')
            plt.close()
        else:
            print("No non-zero values to display in pie chart")

    def generate_report(self, filter_id):
        """生成统计报告"""
        # 获取数据
        issues = self.get_issues(filter_id)
        
        # 如果没有获取到issues，直接返回
        if not issues:
            print("No issues found")
            return
        
        # 成功获取issues后，创建输出目录
        jql = self.jira.filter(filter_id).jql
        target = self.get_target_from_jql(jql)
        self.output_dir = self.create_output_directory(target)
        
        # 分析数据
        df = self.analyze_issues(filter_id)
        
        # 在使用 resolved_issues 之前先定义它
        resolved_issues = []
        all_issues = []
        urgency_details = {
            'U0 Blocking': [], 'U1 Urgent': [],
            'U2 Normal': [], 'U3 Low': [], 'None': []
        }
        
        def convert_to_china_time(date_str):
            """转换UTC时间到中国时区"""
            if not date_str:
                return ''
            try:
                # 解析UTC时间字符串
                utc_time = datetime.strptime(date_str, '%Y-%m-%dT%H:%M:%S.%f%z')
            except ValueError:
                try:
                    # 尝试不带毫秒的格式
                    utc_time = datetime.strptime(date_str, '%Y-%m-%dT%H:%M:%S%z')
                except ValueError:
                    return date_str
                
            # 转换到北京时区
            china_tz = pytz.timezone('Asia/Shanghai')
            china_time = utc_time.astimezone(china_tz)
            return china_time.strftime('%Y-%m-%d %H:%M')  # 格式化为年月日时分
        
        # 获取详细的issue信息并直接转换时间格式
        for issue in issues:
            # 获取 urgency 值
            urgency = str(getattr(issue.fields, 'customfield_11214', 'None'))
            
            # 转换创建时间为中国时区
            created_time = convert_to_china_time(issue.fields.created)
            
            # 准备issue数据
            issue_data = {
                'Issue Key': issue.key,
                'Issue ID': issue.id,
                'Summary': issue.fields.summary,
                'Status': issue.fields.status.name,
                'Urgency': urgency,
                'Technology': getattr(issue.fields, 'customfield_11219', ''),
                'Reporter': getattr(issue.fields.reporter, 'displayName', ''),
                'Assignee': getattr(issue.fields.assignee, 'displayName', ''),
                'Created': created_time,  # 使用转换后的时间
                'Comment': ''
            }
            
            # 添加到相应的列表中
            all_issues.append(issue_data)
            if issue.fields.status.name == 'Resolved':
                resolved_issues.append(issue_data)
            
            # 获取urgency并添加到相应分类
            if urgency in urgency_details:
                urgency_details[urgency].append(issue_data)
        
        # 定义两种列顺序
        columns = [
            'Issue Key', 
            'Issue ID', 
            'Summary',
            'Status',
            'Urgency',  # 添加 Urgency 到列顺序中
            'Technology', 
            'Reporter', 
            'Assignee', 
            'Created'
        ]
        
        # 带Comment的列顺序
        columns_with_comment = columns + ['Comment']
        
        # 修改Excel文件保存路径
        excel_file = os.path.join(self.output_dir, 'bug_statistics.xlsx')
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # 写入Bug Statistics表格
            df.to_excel(writer, sheet_name='Bug Statistics', index=True)
            
            # 获取Bug Statistics工作表
            ws_stats = writer.sheets['Bug Statistics']
            
            # 定义样式
            header_fill = PatternFill(start_color='f8a074', end_color='f8a074', fill_type='solid')  # 淡橙色
            footer_fill = PatternFill(start_color='bff4f9', end_color='bff4f9', fill_type='solid')  # 淡蓝色
            index_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')  # 更淡的橙色
            calibri_font = Font(name='Calibri')
            header_font = Font(name='Calibri', bold=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 为所有数据单元格添加边框和字体
            for row in ws_stats.iter_rows(min_row=1, max_row=ws_stats.max_row, 
                                        min_col=1, max_col=ws_stats.max_column):
                for cell in row:
                    cell.border = thin_border
                    cell.font = calibri_font
            
            # 设置首行样式
            for cell in ws_stats[1]:
                cell.font = header_font
                cell.fill = header_fill
            
            # 设置首列样式（除了头和尾行）
            for row in range(2, ws_stats.max_row+1):
                cell = ws_stats.cell(row=row, column=1)
                cell.fill = index_fill
                cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # 设置尾行样式
            for cell in ws_stats[ws_stats.max_row]:
                cell.fill = footer_fill
                cell.font = header_font
            
            # 写入其他sheet并应用格式
            def format_sheet(worksheet, has_comment=False):
                """格式化工作表"""
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # 设置所有单元格字体为Calibri和边框
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.font = calibri_font
                        cell.border = thin_border
                
                # 设置日期列格式和对齐方式
                date_col = 'I'  # Created 列是第9列
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet[f'{date_col}{row}']
                    if cell.value:
                        cell.alignment = Alignment(horizontal='left')  # 左对齐日期
            
            # 写入 Resolved sheet
            if resolved_issues:
                resolved_df = pd.DataFrame(resolved_issues)
                resolved_df = resolved_df[columns_with_comment]
                resolved_df.to_excel(writer, sheet_name='Resolved', index=False)
                format_sheet(writer.sheets['Resolved'], True)
            
            # 写入 All_bugs sheet
            if all_issues:
                all_bugs_df = pd.DataFrame(all_issues)
                all_bugs_df = all_bugs_df[columns]
                all_bugs_df.to_excel(writer, sheet_name='All_bugs', index=False)
                format_sheet(writer.sheets['All_bugs'], False)
            
            # 写入urgency sheets
            for urgency, issues in urgency_details.items():
                if issues:
                    details_df = pd.DataFrame(issues)
                    if urgency in ['U0 Blocking', 'U1 Urgent']:
                        details_df = details_df[columns_with_comment]
                    else:
                        details_df = details_df[columns]
                    
                    sheet_name = urgency.replace(' ', '_')[:31]
                    details_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    format_sheet(writer.sheets[sheet_name], 
                               urgency in ['U0 Blocking', 'U1 Urgent'])
            
            # 添加时间戳（使用中国时区）
            timestamp = datetime.now(pytz.timezone('Asia/Shanghai')).strftime("%Y-%m-%d %H:%M:%S")
            ws_stats['A1'] = None
        
        print(f"\nExcel report has been saved in '{excel_file}'")
        
        # 检查是否有数据再创建图表
        if not df.empty and df.values.sum() > 0:
            self.create_pie_chart(df.iloc[:-1, :-1])  # 排除总计行和列
            print(f"\nPie chart has been saved in '{os.path.join(self.output_dir, 'bug_status_pie_chart.png')}'")
        else:
            print("\nNo data available for visualization")

if __name__ == "__main__":
    # 创建统计对象
    stats = JiraBugStatistics()
    
    # 使用过滤器ID生成报告
    filter_id = "105508"  # 替换为你的实际过滤器ID
    stats.generate_report(filter_id)